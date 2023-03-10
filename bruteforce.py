from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import combinations


PATH = "data\\actions.xlsx"
CREDIT = 500


def get_data(path):
    """
    Ouvre le fichier xlsx et renvoi une liste d'objets Action
    """
    wb = load_workbook(filename = PATH)
    ws = wb.active

    actions = []
    number = 0
    for row in ws.iter_rows():
        # On saute la ligne d'en tête
        if number > 0:
            name = row[0].value
            cost = row[1].value
            profit = row[2].value
            actions.append(Action(name, cost, profit))
        number += 1
    return actions


class Action:
    def __init__(self, name, cost, profit):
        self.name = name
        # Conversion en centimes pour travailler avec des entiers
        self.cost = int(cost * 100)
        # Valeur en %
        self.profit = profit

    def __str__(self):
        result = str(self.name) + " : " + str(self.cost / 100) + "€ "
        result += str(self.profit) + "% sur 2 ans"
        return result

    @property
    def benefice(self):
        """
        Calcule et renvoi les benefices effectués sur 2 ans
        resultat en centimes
        """
        return int(self.cost * self.profit / 100)



if __name__ == '__main__':
    actions = get_data(PATH)

    # Initialisation de range qui servira pour combinations
    range = len(actions)
    best_benef = 0
    best_combinaison = None
    best_cost = 0
    while range > 1:
        print("Range : " + str(range))
        combinaisons = combinations(actions, range)
        range -= 1
        for combinaison in combinaisons:
            benef_total = 0
            cout_total = 0
            for action in combinaison:
                cout_total += action.cost
                benef_total += action.benefice
            if cout_total <= (CREDIT * 100) and benef_total > best_benef:
                best_benef = benef_total
                best_combinaison = combinaison
                best_cost = cout_total
    if best_combinaison:
        print("Meilleur résultat : ", best_combinaison)
        for e in best_combinaison:
            print(e.name, e.cost / 100, e.benefice / 100)
        print("Cout total : " + str(best_cost / 100))
        print("Benefice total : " + str(best_benef / 100))
    else:
        print("pas de combinaison trouvée")
