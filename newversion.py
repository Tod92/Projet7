from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import combinations, accumulate
from datetime import datetime

PATH = "data2.xlsx"
CREDIT = 500

def chrono_decorator(function):
    def modified(*args,**kwargs):
        a = datetime.now()
        result = function(*args,**kwargs)
        b = datetime.now()
        print("exec time :" + str(b-a))
        return result
    return modified

def get_data(path):
    """
    Ouvre le fichier xlsx et renvoi une liste d'objets Action
    """
    wb = load_workbook(filename = PATH)
    ws = wb.active

    actions = []
    number = 0
    for row in ws.iter_rows():
        # On saute la ligne d'en tête
        if number > 0:
            name = row[0].value
            cost = row[1].value
            profit = row[2].value
            actions.append(Action(name, cost, profit))
        number += 1
    return actions


class Action:
    def __init__(self, name, cost, profit):
        self.name = name
        # Conversion en centimes pour travailler avec des entiers
        self.cost = int(cost * 100)
        self.profit = profit

    def __str__(self):
        result = str(self.name) + " : " + str(self.cost / 100) + "€ "
        result += str(self.profit) + "% sur 2 ans"
        return result

    @property
    def benefice(self):
        """
        Calcule et renvoi les benefices effectués sur 2 ans
        resultat en €
        """
        return int(self.cost * self.profit) / 100

@chrono_decorator
def main():
    actions = get_data(PATH)
    actions.sort(key=lambda x: x.profit,reverse=True)
    accumulated_costs = accumulate([a.cost for a in actions])

    index_in_budget = -1
    budget = CREDIT * 100
    costs_in_bugdet = [cost for cost in accumulated_costs if cost <= budget]
    index_in_budget = len(costs_in_bugdet) - 1

    return actions[:index_in_budget]

if __name__ == '__main__':
    best_combinaison = main()
    cout_total = 0
    benef_total = 0
    for e in best_combinaison:
        print(e.name, e.cost / 100, e.benefice)
        cout_total += e.cost
        benef_total += e.benefice
    print("Cout total : " + str(cout_total))
    print("Benefice total : " + str(benef_total))
